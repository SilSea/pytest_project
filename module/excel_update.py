import openpyxl
import os
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

# กำหนด Path TestSheet
TEST_SHEET_PATH = os.path.realpath('testsheet_lab.xlsx')

# กำหนด Path Logs
LOGS_PATH = os.path.join('logs', 'logs.txt')
SETTING_PATH = os.path.join('logs', 'setting.txt')

# แก้ไขข้อมูลของ Tester กับข้อมูล Environment
def update_name_env():
    # เปิดไฟล์ Excel
    workbook = openpyxl.load_workbook(TEST_SHEET_PATH)

    # เปิดไฟล์ Setting เพื่ออ่านข้อมูล
    with open(SETTING_PATH, "r", encoding="utf-8") as file:
        lines = file.readlines()

    # กำหนดข้อมูล
    tester_name = lines[0].strip().split(": ", 1)[1]
    tester_env = lines[1].strip() + "\n" + lines[2].strip()
    # กำหนดคำค้นหา
    search_dict = {
        "Tester Name": tester_name,
        "Test Environment": tester_env
    }

    # วนลูปทุกชีตในไฟล์
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # วนลูปผ่านแต่ละแถว
        for row in worksheet.iter_rows():
            for cell in row:
                # ถ้าคำในเซลล์ตรงกับคำที่ค้นหา
                if cell.value in search_dict:
                    new_value = search_dict[cell.value]
                    target_cell = cell.offset(0, 1)  # เซลล์ที่อยู่ขวาของคำค้นหา
                    target_cell.value = new_value

    # บันทึกการเปลี่ยนแปลงลงในไฟล์ Excel
    workbook.save(TEST_SHEET_PATH)
    workbook.close()

# อัพเดตข้อมูลการทดสอบลง Excel
def update_excel_with_image(sheet_name, search_word, h_data, img_path):
    # เปิดไฟล์ Excel
    workbook = openpyxl.load_workbook(TEST_SHEET_PATH)

    # เลือก sheet ที่ต้องการทำงาน
    sheet = workbook[sheet_name]

    # อ่านข้อความใน Logs ไฟล์
    with open(LOGS_PATH, 'r', encoding='utf-8') as file:
        g_data = file.read()

    # วนลูปผ่านแต่ละแถว
    for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row): 
        for cell in row:
            # ถ้าคำในเซลล์ตรงกับคำที่ค้นหา
            if cell.value == search_word:

                # ตรวจสอบ PASS/FAIL และเปลี่ยนสีข้อความในเซลล์ H
                if h_data == "PASS":
                    # เขียนข้อมูลใหม่ลงในเซลล์ Actual (G)
                    sheet[f'G{cell.row}'] = sheet[f'F{cell.row}'].value
                    sheet[f'G{cell.row}'].font = Font(color="000000")  # สีดำ
                    # เขียนข้อมูลในเซลล์ H พร้อมกับเปลี่ยนสีข้อความเป็นสีเขียว
                    sheet[f'H{cell.row}'] = h_data
                    sheet[f'H{cell.row}'].font = Font(color="2d8a13")  # สีเขียว (ใช้รหัส RGB)
                elif h_data == "FAILED":
                    # เขียนข้อมูลใหม่ลงในเซลล์ Actual (G)
                    sheet[f'G{cell.row}'] = g_data
                    sheet[f'G{cell.row}'].font = Font(color="000000")  # สีดำ
                    # เขียนข้อมูลในเซลล์ H พร้อมกับเปลี่ยนสีข้อความเป็นสีแดง
                    sheet[f'H{cell.row}'] = h_data
                    sheet[f'H{cell.row}'].font = Font(color="FF0000")  # สีแดง (ใช้รหัส RGB)

                # ใส่รูปในเซลล์ I
                img = Image(img_path)  # โหลดรูปภาพ

                # คำนวณขนาดของเซลล์ I
                col_width = sheet.column_dimensions['I'].width  # ความกว้างของเซลล์ I
                row_height = sheet.row_dimensions[cell.row].height  # ความสูงของเซลล์ที่พบ

                # ปรับขนาดรูปให้เท่ากับขนาดของเซลล์
                img.width = col_width * 8
                img.height = row_height * 1.3

                sheet.add_image(img, f'I{cell.row}')  # ใส่รูปในเซลล์ I ที่แถวที่พบคำค้นหา

                # บันทึกการเปลี่ยนแปลงในไฟล์
                workbook.save(TEST_SHEET_PATH)
                break  # หยุดค้นหาต่อจากแถวนี้
